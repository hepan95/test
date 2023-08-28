#!/usr/bin/env/python
# --coding:utf-8--
# @fileName: common_funtion.py.py
# @Time:2022/10/18 11:07
# @Author:PH
import  socket  #获取本地主机IP地址
import datetime  #获取时间模块
import os        #获取主机IP地址模块
import string  #自动生成含税模块
import random
import time

import openpyxl
import xlrd
# from imbox import Imbox
# from datetime import datetime

class Common_page():
    def __init__(slif):
        pass
    def projects_path(self):    #查看模块路径！！！！！！！兼容性强 >>项目地址
        current_path=os.path.abspath(__file__)
        father_path=os.path.abspath(os.path.dirname(current_path)+os.path.sep)
        new_path=os.path.abspath(os.path.dirname(father_path)+os.path.sep)
        # print(new_path)
        return new_path
    '''获取内网IP地址'''
    def get_id01(self):
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        # print(ip)
        return ip
    def get_today(self):     #获取今天的年月日的类方法
        now_time=datetime.datetime.now().strftime("%Y-%m-%d")
        # print(now_time)
        return  now_time
    def get_today001(self):
        # 获取今天（现在时间）
        today = datetime.datetime.today()
        yesterday = today - datetime.timedelta(days=1) # 昨天
        tomorrow = today + datetime.timedelta(days=1)  # 明天
        date = datetime.date.today()  # 获取当前日期
        y = today + datetime.timedelta(days=5) # 获取的时间---=1代表1天-
        z = today + datetime.timedelta(days=10) #
        q = today + datetime.timedelta(days=20)  #
        c = today + datetime.timedelta(days=50)  #
        t  = today + datetime.timedelta(days=100)  #
        w = today + datetime.timedelta(days=200)  #
        n = today + datetime.timedelta(days=0)  #
        p = today + datetime.timedelta(days=5)
        h = y.strftime('%Y-%m-%d')   #今天
        y = y.strftime('%Y-%m-%d %H:%M:%S')  # +5天
        z = z.strftime('%Y-%m-%d %H:%M:%S')  # +10天
        q = q.strftime('%Y-%m-%d %H:%M:%S')  #+ 20天
        c = c.strftime('%Y-%m-%d ')  #+50天
        t = t.strftime('%Y-%m-%d ')  #+100天
        w = w.strftime('%Y-%m-%d ')  # + 200天
        g = n.strftime('%Y%m%d%H%M%S')  #+ 20天
        k = n.strftime('%Y-%m-%d %H:%M:%S')  #今天 时分秒
        j = n.strftime('%Y-%m-%d' )  #今天
        f = n.strftime('%Y%m%d%H%M%S')  #今天 时分秒
        f = f[2:14]   #切片
        ll = p.strftime('%Y-%m-%d' )   #+5天
        zz = n.strftime("%H%M")
        cc = p.strftime('%Y-%m-%d %H:%M:%S' )   #+5天
        # print(cc)
        # print(today)
        # print(z)
        # print(q)
        # print(h)
        # print(c)
        # print(t)
        # print(w)
        # print(g)
        # print(k)
        # print(j)
        # print(f)
        # cc= int(date.strftime('%Y%m%d'))
        # print(cc)
        # print(ll)
        # print(f[2:14])

        return today,y,z,q,h,c,t,w,g,h,k,j,f,ll,zz,cc

    def start(self):

        x =  ''.join(random.sample(string.digits,10))

        # print(x)
        return x

    '''excel写入数据'''
    def excel001(self,time1,ip_zd,bm1):
        # 解题思路：遍历表，判断每一行取出字典中，存在则更新对应key；不存在，则跳过这一行。
        # ip_zd = Common_page().projects_path()
        # bm= Common_page().get_today001()[8]
        # bm1 = 'AL0-T'+bm
        wb = openpyxl.load_workbook(r'{}\Test_data\FBA.xlsx'.format(ip_zd))
        ws = wb.active  # 打开表格
        price_update = {'AL0 NO. （必填）': bm1}
        price_update1 = {'预约到达工厂日期（必填）': time1}
        for rowNum in range(2, ws.max_row + 1):  # 指的是 第二列，第二行
            # 取出表单中每一行第一列的名称
            produce = ws.cell(row=rowNum, column=1).value  # 遍历表格key
            # print(produce)
            # 判断名称是否存在于待更新的字典中
            if produce in price_update.keys():  # 如果key = price_update  ，则写入内容
                ws.cell(row=rowNum, column=2).value = price_update[produce]
                pass

            # 判断名称是否存在于待更新的字典中
            if produce in price_update1.keys():  # 如果key = price_update1  ，则写入内容
                ws.cell(row=rowNum, column=2).value = price_update1[produce]
                pass
        # 将更新后的文件的另存为，覆盖原文件数据
        wb.save(r'{}\Test_data\FBA.xlsx'.format(ip_zd))
        # print(wb)
        # 查看更新后的xlsx更新成功。

    '''获取星期'''
    def Week(self,time="2023-08-13"):

        date_str = time
        date_format = '%Y-%m-%d'
        date = datetime.datetime.strptime(date_str, date_format)

        weekday = date.weekday()+1
        # print("2023年8月9日是星期", weekday)
        return weekday

    '''存储附件'''
    def save_attachments(self,attachments):
        # 设置存放路径
        save_dir = r"{}\Test_data\inquiry_document".format(Common_page().projects_path())
        if not os.path.exists(save_dir):
            os.mkdir(save_dir)
        fujian_path_list = []
        for attachment in attachments:
            save_path = os.path.join(save_dir, attachment['filename'])
            with open(save_path, 'wb') as f:
                f.write(attachment['content'].getvalue())
            fujian_path_list.append(save_path)
            # print(fujian_path_list)
        return fujian_path_list

    # '''qq邮箱获取，下载附件'''
    # def doMain(self,id='1044726725@qq.com', password="fndwnamwwtlgbejf",xj_dh="AL0-T202308111517"):
    #     #  Imbox(服务器,邮箱用户名,密码,SSL加密) password 是认证码，非密码：由QQ邮箱服务生成
    #     with Imbox('imap.qq.com', id, password, ssl=True) as imbox:
    #         all_inbox_messages = imbox.messages()  # 获取全部邮件
    #         # email_list = email_server.messages(unread=True)  # 未读邮件
    #         # read_inbox_messages = self.email_server.messages(unread=False)  # 已读邮件
    #         # flagged_inbox_messages = self.email_server.messages(flagged=True)  # 红旗标记邮件
    #         i = 0
    #         for uid, message in all_inbox_messages:
    #             i += 1
    #             if i > 150:
    #                 print('test finish')
    #                 break
    #             # imbox.mark_seen(uid)   # 标记为已读
    #             # imbox.delete(uid)  # 删除
    #             # print(message.subject)  # 邮件主题
    #             # print(message.sent_from)  # 发件人
    #             # print(message.sent_to)  # 收件人
    #             # print(message.date)  # 发送日期
    #             # print(message.body['plain']) # 邮件文本格式正文
    #             # print(message.body['html']) # html格式正文
    #
    #             if message.subject == "{}询价结果".format(xj_dh):  # 如果邮件主题等于
    #                 Common_page().save_attachments(message.attachments)

    '''读取excle表格'''
    import warnings
    warnings.filterwarnings('ignore')
    def read_excel001(self,save_dir='D:\\work\\xiangmu\\test_ICT_api'):   #表格路径，能直接打开表格
        save_dir = r'{}'.format(save_dir)
        csdata = "{}".format(save_dir)
        xls = openpyxl.load_workbook(csdata)
        niubi = xls["Sheet0"]  # 读取表格的子表格名字
        # print(niubi.max_row)  # 读最大行
        # print(niubi.max_column)  # 读最大列
        max_row = niubi.max_row
        max_column = niubi.max_column
        list_a = []
        for x in range(2, max_row + 1):  # 从表格的第二行开始读取
            dict = {}
            # for y in range(1, max_column + 1):
            for y in range(1, 45):
                key = niubi.cell(row=1, column=y).value  #
                # print(key)
                value = niubi.cell(row=x, column=y).value  #
                dict[key] = value
            # print(dict)
            list_a.append(dict)
        # print(list_a)
        # Expected_pickup_date=list_a[0]["期望提货日期(Expected Pickup Date)"]
        # print(Expected_pickup_date)
        # Expected_pickup_time = list_a[0]["期望提货时间(Expected Pickup Time)"]
        # print(Expected_pickup_time)



        return list_a

    '''excel写入数据'''
    def read_excel(self,file_path,ALO,time1):
        import xlrd, xlwt
        from xlutils.copy import copy
        # 读取文件
        # file_path = r"D:\work\xiangmu\test_ICT_api\Test_data\FBA_inquiry_list.xls"
        read_file = xlrd.open_workbook(file_path, formatting_info=True)
        # 参数注释：
        # file_path：文件路径，包含文件的全名称
        # formatting_info=True：保留Excel的原格式
        # 将文件复制到内存
        write_data = copy(read_file)
        # 读取复制后文件的sheet1
        write_save = write_data.get_sheet(0)
        # 写入数据
        write_save.write(1, 1, ALO )  #询价单号'AL0-T230808095141'
        write_save.write(3, 1, time1)   #提货时间  年月日 2023-08-11
        # 参数注释：
        # x,y:写入目标格的位置坐标
        # value：写入数据
        # 保存写入数据后的文件到原文件路径
        write_data.save(file_path)
        # print(write_data)
    #
    #     '''qq邮箱获取，下载附件'''
    #
    # def doMain1(self, id='hepan@hercules-logistics.com', password="j43zFLnpzAYHB8LU", xj_dh="AL0-T202308111517"):
    #     #  Imbox(服务器,邮箱用户名,密码,SSL加密) password 是认证码，非密码：由QQ邮箱服务生成
    #     with Imbox('imap.qq.com', id, password, ssl=True) as imbox:
    #         all_inbox_messages = imbox.messages()  # 获取全部邮件
    #         # email_list = email_server.messages(unread=True)  # 未读邮件
    #         # read_inbox_messages = self.email_server.messages(unread=False)  # 已读邮件
    #         # flagged_inbox_messages = self.email_server.messages(flagged=True)  # 红旗标记邮件
    #         i = 0
    #         for uid, message in all_inbox_messages:
    #             i += 1
    #             if i > 150:
    #                 print('test finish')
    #                 break
    #             # imbox.mark_seen(uid)   # 标记为已读
    #             # imbox.delete(uid)  # 删除
    #             print(message.subject)  # 邮件主题
    #             # print(message.sent_from)  # 发件人
    #             # print(message.sent_to)  # 收件人
    #             # print(message.date)  # 发送日期
    #             # print(message.body['plain']) # 邮件文本格式正文
    #             # print(message.body['html']) # html格式正文
    #
    #             if message.subject == "{}询价结果".format(xj_dh):  # 如果邮件主题等于
    #                 Common_page().save_attachments(message.attachments)

if __name__ == '__main__':
    run=Common_page()
    # run.get_today001()
    # run.read_excel001()
    # run.read_excel()
    # run.read_excel()

